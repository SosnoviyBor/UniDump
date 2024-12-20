import expression.associator as associator
import expression.comutator as comutator
import expression.evaluator as evaluator
import expression.tree as tree


import openpyxl as xl


def test(baseExp:str):
    results = {}
    processes = 6
    
    # # graph rendering snippet
    # t.generateGraph()
    # graph = t.generateGraph()
    # graph.render(filename='graph.dot', directory='generated_graphs', view=True, format='png') 
    
    # base
    t = tree.parseExpression(baseExp)
    speed, efficiency = evaluator.evaluate(t, processes, False)
    results["base"] = {}
    results["base"]["exp"] = baseExp
    results["base"]["speed"] = speed
    results["base"]["efficiency"] = efficiency
    
    # associated
    exp = associator.associate(baseExp, False)
    t = tree.parseExpression(exp)
    speed, efficiency = evaluator.evaluate(t, processes, False)
    results["ass"] = {}
    results["ass"]["exp"] = exp
    results["ass"]["speed"] = speed
    results["ass"]["efficiency"] = efficiency
    
    # 5 commutated
    gen = comutator.comutate(baseExp, False)
    results["com"] = {"exp": [], "speed":[], "efficiency": []}
    for _ in range(10):
        exp = next(gen)
        if not exp:
            break
        t = tree.parseExpression(exp)
        speed, efficiency = evaluator.evaluate(t, processes, False)
        results["com"]["exp"].append(exp)
        results["com"]["speed"].append(speed)
        results["com"]["efficiency"].append(efficiency)
    
    print("Виміри завершені")
    
    # writing the results
    wb = xl.Workbook()
    ws = wb.active
    
    ws["A1"] = "Expression"
    ws["B1"] = "Eval speed"
    ws["C1"] = "Speedup"
    ws["D1"] = "Efficiency"
    
    i = 2
    for item in results.values():
        if type(item["exp"]) is str:
            ws[f"A{i}"] = item["exp"]
            ws[f"B{i}"] = item["speed"]
            ws[f"C{i}"] = round(results["base"]["speed"] / item["speed"], 3)
            ws[f"D{i}"] = item["efficiency"]
            i += 1
        
        elif type(item["exp"]) is list:
            for j in range(len(results["com"]["exp"])):
                ws[f"A{i}"] = item["exp"][j]
                ws[f"B{i}"] = item["speed"][j]
                ws[f"C{i}"] = round(results["base"]["speed"] / item["speed"][j], 3)
                ws[f"D{i}"] = item["efficiency"][j]
                i += 1
        i += 1
    
    wb.save("lab6.xlsx")
    print("Результати записані")